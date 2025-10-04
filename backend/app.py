"""Minimal backend API for Acuzen Autofill (Render/Fly/Railway deployable).

Routes:
  - POST /api/upload (base64 CSV/XLSX -> normalized records)
  - POST /api/cases (store record)
  - GET  /api/cases (list stored)
  - POST /api/reset (clear DB)
  - GET  /healthz (health)

Env:
  - PORT (default 8000)
  - DB_PATH (default ./data/mock_cases.db)
  - ALLOWED_ORIGINS (comma-separated, default *)
"""
from __future__ import annotations

import base64
import binascii
import csv
import json
import os
import sqlite3
import tempfile
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from typing import Any, Dict, Iterable, List


ROOT_DIR = Path(__file__).resolve().parent
DATA_DIR = ROOT_DIR / "data"
DATA_DIR.mkdir(parents=True, exist_ok=True)
DB_PATH = Path(os.getenv("DB_PATH", str(DATA_DIR / "mock_cases.db")))
ALLOWED = [o.strip() for o in os.getenv("ALLOWED_ORIGINS", "*").split(",") if o.strip()] or ["*"]


def init_db() -> None:
    with sqlite3.connect(DB_PATH) as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cases (
                case_id TEXT PRIMARY KEY,
                reaction_reported_term TEXT,
                meddra_level TEXT,
                meddra_term_text TEXT,
                meddra_code TEXT,
                meddra_version TEXT,
                onset_date TEXT,
                seriousness TEXT,
                suspect_drug TEXT,
                dose_text TEXT,
                outcome TEXT,
                narrative TEXT,
                raw_payload TEXT
            )
            """
        )
        conn.commit()


def parse_csv_bytes(data: bytes) -> List[Dict[str, str]]:
    text = data.decode("utf-8-sig")
    reader = csv.DictReader(text.splitlines())
    return [normalize_row(row) for row in reader]


def parse_xlsx_bytes(data: bytes) -> List[Dict[str, str]]:
    from openpyxl import load_workbook  # optional dep declared in requirements

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(data)
        tmp_path = Path(tmp.name)
    try:
        wb = load_workbook(filename=str(tmp_path), read_only=True)
        sheet = wb.active
        header: List[str] = []
        rows: List[Dict[str, str]] = []
        for r_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if r_idx == 1:
                header = [str(c).strip() if c is not None else "" for c in row]
                continue
            record = {header[i]: (str(c).strip() if c is not None else "") for i, c in enumerate(row) if header[i]}
            if any(record.values()):
                rows.append(normalize_row(record))
        wb.close()
        return rows
    finally:
        tmp_path.unlink(missing_ok=True)


def normalize_row(row: Dict[str, str]) -> Dict[str, str]:
    def get(*keys: str) -> str:
        for k in keys:
            v = row.get(k)
            if v:
                return str(v).strip()
        return ""

    reaction = get("reaction_reported_term", "reaction_term", "reported_term")
    return {
        "case_id": get("case_id"),
        "reaction_reported_term": reaction,
        "meddra_level": get("meddra_level") or "PT",
        "meddra_term_text": get("meddra_term_text", "meddra_text", "reaction_term") or reaction,
        "meddra_code": get("meddra_code", "reaction_code", "llt_code", "pt_code"),
        "meddra_version": get("meddra_version") or "MOCK-1.0",
        "onset_date": get("onset_date", "reaction_onset_date"),
        "seriousness": ("Serious" if get("seriousness", "serious").lower().startswith("serious") else (get("seriousness") or "Non-serious")),
        "suspect_drug": get("suspect_drug", "drug_name", "medicinal_product"),
        "dose_text": get("dose_text", "dose"),
        "outcome": get("outcome", "reaction_outcome"),
        "narrative": get("narrative", "case_narrative", "summary", "indication"),
        "raw_payload": json.dumps(row, ensure_ascii=False),
    }


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, directory: str | None = None, **kwargs: Any) -> None:
        directory = directory or str(ROOT_DIR)
        super().__init__(*args, directory=directory, **kwargs)

    def do_GET(self) -> None:  # noqa: N802
        if self.path == "/healthz":
            self._send_json({"ok": True})
            return
        if self.path == "/api/cases":
            with sqlite3.connect(DB_PATH) as conn:
                conn.row_factory = sqlite3.Row
                rows = conn.execute("SELECT * FROM cases ORDER BY case_id").fetchall()
            self._send_json([dict(r) for r in rows])
            return
        if self.path == "/":
            # 이 백엔드는 UI를 제공하지 않습니다.
            self._send_json({"service": "acuzen-autofill-api", "ok": True})
            return
        super().do_GET()

    def do_OPTIONS(self) -> None:  # noqa: N802
        if self.path.startswith("/api/") or self.path == "/healthz":
            self.send_response(HTTPStatus.NO_CONTENT)
            self._write_cors()
            self.end_headers()
            return
        super().do_OPTIONS()

    def do_POST(self) -> None:  # noqa: N802
        if self.path == "/api/upload":
            self._upload()
            return
        if self.path == "/api/cases":
            self._create_case()
            return
        if self.path == "/api/reset":
            self._reset()
            return
        self.send_error(HTTPStatus.NOT_FOUND, "Unsupported endpoint")

    # ---- handlers -----------------------------------------------------
    def _upload(self) -> None:
        length = int(self.headers.get("Content-Length", 0))
        payload = self.rfile.read(length)
        try:
            data = json.loads(payload.decode("utf-8"))
        except json.JSONDecodeError:
            self.send_error(HTTPStatus.BAD_REQUEST, "Invalid JSON")
            return

        filename = str(data.get("filename", "")).strip()
        content = data.get("content")
        if not filename or not content:
            self.send_error(HTTPStatus.BAD_REQUEST, "Missing filename or content")
            return

        base64_data = content.split(",", 1)[-1]
        try:
            file_bytes = base64.b64decode(base64_data)
        except (ValueError, binascii.Error) as exc:
            self.send_error(HTTPStatus.BAD_REQUEST, f"Invalid base64: {exc}")
            return

        suffix = Path(filename).suffix.lower()
        if suffix == ".csv":
            records = parse_csv_bytes(file_bytes)
        elif suffix in (".xlsx", ".xls"):
            records = parse_xlsx_bytes(file_bytes)
        else:
            self.send_error(HTTPStatus.BAD_REQUEST, f"Unsupported file type: {suffix}")
            return

        res = {
            "filename": filename,
            "count": len(records),
            "records": records[:200],
        }
        self._send_json(res)

    def _create_case(self) -> None:
        length = int(self.headers.get("Content-Length", 0))
        payload = self.rfile.read(length)
        try:
            data = json.loads(payload.decode("utf-8"))
        except json.JSONDecodeError:
            self.send_error(HTTPStatus.BAD_REQUEST, "Invalid JSON")
            return

        record = {
            "case_id": str(data.get("case_id", "")).strip(),
            "reaction_reported_term": data.get("reaction_reported_term"),
            "meddra_level": data.get("meddra_level"),
            "meddra_term_text": data.get("meddra_term_text"),
            "meddra_code": data.get("meddra_code"),
            "meddra_version": data.get("meddra_version"),
            "onset_date": data.get("onset_date"),
            "seriousness": data.get("seriousness"),
            "suspect_drug": data.get("suspect_drug"),
            "dose_text": data.get("dose_text"),
            "outcome": data.get("outcome"),
            "narrative": data.get("narrative"),
            "raw_payload": json.dumps(data.get("raw_payload", data), ensure_ascii=False),
        }
        try:
            with sqlite3.connect(DB_PATH) as conn:
                conn.execute(
                    """
                    INSERT OR REPLACE INTO cases (
                      case_id, reaction_reported_term, meddra_level, meddra_term_text, meddra_code,
                      meddra_version, onset_date, seriousness, suspect_drug, dose_text, outcome, narrative, raw_payload
                    ) VALUES (
                      :case_id, :reaction_reported_term, :meddra_level, :meddra_term_text, :meddra_code,
                      :meddra_version, :onset_date, :seriousness, :suspect_drug, :dose_text, :outcome, :narrative, :raw_payload
                    )
                    """,
                    record,
                )
                conn.commit()
        except sqlite3.DatabaseError as exc:
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR, f"DB error: {exc}")
            return
        self._send_json({"status": "saved"}, status=HTTPStatus.CREATED)

    def _reset(self) -> None:
        try:
            with sqlite3.connect(DB_PATH) as conn:
                conn.execute("DELETE FROM cases")
                conn.commit()
        except sqlite3.DatabaseError as exc:
            self.send_error(HTTPStatus.INTERNAL_SERVER_ERROR, f"DB reset error: {exc}")
            return
        self._send_json({"status": "reset"})

    # ---- utils --------------------------------------------------------
    def _allowed_origin(self) -> str:
        origin = self.headers.get("Origin")
        if not origin:
            return "*" if "*" in ALLOWED else ALLOWED[0]
        if "*" in ALLOWED or origin in ALLOWED:
            return origin
        # fallback: deny by echoing none (browsers will block)
        return "null"

    def _write_cors(self) -> None:
        origin = self._allowed_origin()
        self.send_header("Access-Control-Allow-Origin", origin)
        self.send_header("Vary", "Origin")
        self.send_header("Access-Control-Allow-Methods", "GET,POST,OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type, Authorization")

    def _send_json(self, obj: Any, status: int = 200) -> None:
        self.send_response(status)
        self._write_cors()
        self.send_header("Content-Type", "application/json")
        self.end_headers()
        self.wfile.write(json.dumps(obj, ensure_ascii=False).encode("utf-8"))


def main() -> int:
    init_db()
    port = int(os.getenv("PORT", "8000"))
    httpd = ThreadingHTTPServer(("", port), Handler)
    print(f"Serving on http://0.0.0.0:{port}")
    try:
        httpd.serve_forever()
    finally:
        httpd.server_close()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

