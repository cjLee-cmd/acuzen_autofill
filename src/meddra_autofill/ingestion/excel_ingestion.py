"""Excel/CSV ingestion utilities."""
from __future__ import annotations

import csv
from pathlib import Path
from typing import Iterable, Iterator, List

from ..models import CaseRecord, records_from_iterable


class ExcelIngestor:
    """Loads case records from CSV or Excel files."""

    SUPPORTED_EXTENSIONS = {".csv", ".xlsx", ".xls"}

    def __init__(self, encoding: str = "utf-8-sig") -> None:
        self.encoding = encoding

    def load_records(self, file_path: str | Path) -> List[CaseRecord]:
        rows = self.load_rows(file_path)
        return records_from_iterable(rows)

    def load_rows(self, file_path: str | Path) -> List[dict[str, str]]:
        path = Path(file_path)
        if path.suffix.lower() not in self.SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"Unsupported file type '{path.suffix}'. Expected one of {self.SUPPORTED_EXTENSIONS}."
            )

        if path.suffix.lower() == ".csv":
            return list(self._read_csv(path))
        return list(self._read_excel(path))

    def _read_csv(self, path: Path) -> Iterator[dict[str, str]]:
        with path.open("r", encoding=self.encoding, newline="") as handle:
            reader = csv.DictReader(handle)
            for row in reader:
                yield {k.strip(): v.strip() if isinstance(v, str) else v for k, v in row.items()}

    def _read_excel(self, path: Path) -> Iterator[dict[str, str]]:
        try:
            from openpyxl import load_workbook
        except ImportError as exc:  # pragma: no cover - optional dependency
            raise RuntimeError(
                "openpyxl is required to read Excel files. Install it with 'pip install openpyxl'."
            ) from exc

        workbook = load_workbook(filename=path, read_only=True)
        sheet = workbook.active
        header: List[str] = []
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row_idx == 1:
                header = [str(cell).strip() if cell is not None else "" for cell in row]
                continue
            record = {header[idx]: (str(cell).strip() if cell is not None else "") for idx, cell in enumerate(row) if header[idx]}
            if any(value for value in record.values()):
                yield record
        workbook.close()


__all__ = ["ExcelIngestor"]
